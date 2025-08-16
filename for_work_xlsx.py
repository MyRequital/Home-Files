import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import re


DATA_FOLDER = Path("tables")
DATA_FOLDER.mkdir(exist_ok=True)

VALID_CELL_REGEX = re.compile(r"^[A-Z]+[1-9][0-9]*$")  # Пример: A1, B12, AA100

def is_valid_cell_address(address: str) -> bool:
    return VALID_CELL_REGEX.fullmatch(address) is not None


def create_table(table_name: str, user_id: str) -> None:
    """
    Создаёт новую таблицу Excel и позволяет пользователю вручную заполнить ячейки.
    """
    user_folder = DATA_FOLDER / user_id
    user_folder.mkdir(exist_ok=True)
    file_path = user_folder / f"{table_name}.xlsx"

    if file_path.exists():
        print("❌ Таблица уже существует.")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    print("▶ Ввод данных (введите '0' как адрес ячейки для завершения)\n")

    try:
        while True:
            cell_address = input("Введите ячейку (например, A3): ").strip().upper()

            if cell_address == "0":
                break

            if not is_valid_cell_address(cell_address):
                print("⚠️ Адрес ячейки должен быть латиницей и в формате вроде 'A1'. Попробуй снова.")
                continue

            cell_value = input("Чем заполнить: ")
            sheet[cell_address] = cell_value

        workbook.save(file_path)
        print(f"✅ Таблица '{table_name}' сохранена для пользователя '{user_id}' по пути: {file_path}")

    except Exception as error:
        print(f"❌ Ошибка при создании таблицы: {error}")


def work_with_table(table_name: str, user_id: str) -> None:
    """
    Позволяет пользователю добавлять или изменять значения ячеек в таблице.
    """
    file_path = DATA_FOLDER / user_id / f"{table_name}.xlsx"

    if not file_path.exists():
        print("❌ Таблица не найдена.")
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    print("▶ Режим редактирования таблицы (введите '0' для выхода)\n")

    try:
        while True:
            cell_address = input("Введите ячейку (например, A3): ").strip().upper()
            if cell_address == "0":
                break

            if not is_valid_cell_address(cell_address):
                print("⚠️ Некорректный адрес ячейки. Используйте латинские буквы и номер, например 'B2'.")
                continue

            current_value = sheet[cell_address].value
            if current_value is not None:
                print(f"⚠️ Ячейка ({cell_address}) уже содержит: '{current_value}'")
                choice = input("Заменить значение? (1 - Да, 0 - Нет): ").strip()
                if choice != "1":
                    continue

            new_value = input("Новое значение: ")
            sheet[cell_address] = new_value

        workbook.save(file_path)
        print("✅ Изменения сохранены.")

    except Exception as error:
        print(f"❌ Ошибка при работе с таблицей: {error}")


def show_table(table_name: str, user_id: str) -> None:
    """
    Показывает содержимое таблицы пользователя в формате (A1) Значение.
    """
    file_path = DATA_FOLDER / user_id / f"{table_name}.xlsx"

    if not file_path.exists():
        print("❌ Таблица не найдена.")
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    print("\n📊 Содержимое таблицы:")

    output_lines = []

    for row in sheet.iter_rows():
        line = []
        for cell in row:
            if cell.value is not None:
                line.append(f"({cell.coordinate}) {cell.value}")
        if line:
            output_lines.append(" | ".join(line))

    if output_lines:
        print("\n".join(output_lines))
    else:
        print("🕳️ Таблица пуста.")



def drop_table(table_name: str, user_id: str) -> None:
    """
    Удаляет таблицу пользователя.
    """
    file_path = DATA_FOLDER / user_id / f"{table_name}.xlsx"

    if not file_path.exists():
        print("❌ Таблица не найдена.")
        return

    file_path.unlink()
    print(f"🗑️ Таблица '{table_name}' удалена.")


def main() -> None:
    print(
        f"\n1 - Создать таблицу (Название)"
        f"\n2 - Работать с таблицей (Добавить строку/(-и))"
        f"\n3 - Показать таблицу"
        f"\n4 - Удалить таблицу\n"
    )

    try:
        choice = int(input("Выберите действие: "))

        if choice in {1, 2, 3, 4}:
            table_name = input("Название таблицы: ")
            user_id = input("ID пользователя: ")

        match choice:
            case 1:
                create_table(table_name, user_id)

            case 2:
                work_with_table(table_name, user_id)

            case 3:
                show_table(table_name, user_id)

            case 4:
                drop_table(table_name, user_id)

            case _:
                print("❓ Неизвестная команда")


    except Exception as error:

        print(f"❌ Ошибка выполнения: {error}")


while True:
    main()